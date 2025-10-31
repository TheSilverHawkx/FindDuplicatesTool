from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

_MB_divider = 1024**2

def write_report(dups_by_hash: dict[str, tuple[Path,int]], output_path: Path) -> int:
    if output_path.exists() and output_path.is_dir():
        raise IsADirectoryError(output_path)

    # keep only real duplicates
    duplicates = {h: ps for h, ps in dups_by_hash.items() if len(ps) > 1}

    # always write an .xlsx file
    if output_path.suffix.lower() != ".xlsx":
        output_path = output_path.with_suffix(".xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "duplicates"

    # header
    ws.append(["checksum", "file_path","size"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    # rows (Unicode—including Cyrillic—is fully supported)
    count_rows = 0
    for checksum in sorted(duplicates.keys()):
        # sort by path (string form, case-insensitive)
        for path_obj, size in sorted(duplicates[checksum], key=lambda x: str(x[0]).lower()):
            ws.append([checksum, str(path_obj), f"{round(path_obj.stat().st_size / _MB_divider,2)} MB"])
            count_rows += 1


    # autosize columns
    for col_idx in range(1, 4):  # now we have 3 columns
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0)
            for c in ws[col_letter]
        )
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 120)

    wb.save(output_path)
    return count_rows
